import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import json

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─── ЦВЕТА И СТИЛЬ ───────────────────────────────────────────────────────────
BG_DARK      = "#0F1923"
BG_CARD      = "#16232E"
BG_INPUT     = "#1C2E3D"
BG_ROW_ODD   = "#162130"
BG_ROW_EVEN  = "#1A2840"
ACCENT       = "#00C896"
ACCENT_DARK  = "#009E78"
ACCENT2      = "#F0A500"
TEXT_MAIN    = "#E8F4F0"
TEXT_DIM     = "#7A9FAF"
TEXT_HEADER  = "#FFFFFF"
RED          = "#FF5A5A"
BORDER_CLR   = "#1E3448"

FONT_TITLE   = ("Segoe UI", 22, "bold")
FONT_HEAD    = ("Segoe UI", 11, "bold")
FONT_NORM    = ("Segoe UI", 10)
FONT_SMALL   = ("Segoe UI", 9)
FONT_BTN     = ("Segoe UI", 10, "bold")
FONT_BIG     = ("Segoe UI", 26, "bold")

DATA_FILE = "data.json"

# ─── ХРАНИЛИЩЕ ДАННЫХ ────────────────────────────────────────────────────────
def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {"prikhod": [], "raskhod": []}

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ─── ГЛАВНОЕ ПРИЛОЖЕНИЕ ───────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("📦 База Досок — Складской учёт")
        self.geometry("1100x700")
        self.minsize(900, 600)
        self.configure(bg=BG_DARK)
        self.data = load_data()

        self.update_idletasks()
        x = (self.winfo_screenwidth() - 1100) // 2
        y = (self.winfo_screenheight() - 700) // 2
        self.geometry(f"1100x700+{x}+{y}")

        self._build_ui()

    def _build_ui(self):
        nav = tk.Frame(self, bg=BG_CARD, width=220)
        nav.pack(side="left", fill="y")
        nav.pack_propagate(False)

        logo_frame = tk.Frame(nav, bg=BG_CARD)
        logo_frame.pack(fill="x", pady=(28, 8))
        tk.Label(logo_frame, text="📦", font=("Segoe UI", 32),
                 bg=BG_CARD, fg=ACCENT).pack()
        tk.Label(logo_frame, text="База Досок", font=FONT_TITLE,
                 bg=BG_CARD, fg=TEXT_HEADER).pack()
        tk.Label(logo_frame, text="Складской учёт", font=FONT_SMALL,
                 bg=BG_CARD, fg=TEXT_DIM).pack(pady=(2, 0))

        tk.Frame(nav, bg=BORDER_CLR, height=1).pack(fill="x", padx=20, pady=20)

        self.nav_buttons = {}
        menu_items = [
            ("📊", "Дашборд",   "dashboard"),
            ("📥", "Приход",    "prikhod"),
            ("📤", "Расход",    "raskhod"),
            ("📋", "Отчёт",     "otchet"),
        ]
        for icon, label, key in menu_items:
            btn = NavButton(nav, icon, label, lambda k=key: self.show_page(k))
            btn.pack(fill="x", padx=12, pady=3)
            self.nav_buttons[key] = btn

        tk.Frame(nav, bg=BORDER_CLR, height=1).pack(fill="x", padx=20, pady=20)
        export_btn = tk.Button(nav, text="💾  Экспорт в Excel",
                               font=FONT_BTN, bg=ACCENT2, fg="#000",
                               relief="flat", cursor="hand2",
                               activebackground="#D4920A", activeforeground="#000",
                               command=self.export_excel, padx=10, pady=8)
        export_btn.pack(fill="x", padx=12, pady=4)

        del_all_btn = tk.Button(nav, text="🗑  Удалить всё",
                               font=FONT_BTN, bg="#3A1A1A", fg=RED,
                               relief="flat", cursor="hand2",
                               activebackground="#5A2020", activeforeground=RED,
                               command=self._confirm_delete_all, padx=10, pady=8)
        del_all_btn.pack(fill="x", padx=12, pady=(0,16))

        self.content = tk.Frame(self, bg=BG_DARK)
        self.content.pack(side="left", fill="both", expand=True)

        self.pages = {}
        self.pages["dashboard"] = DashboardPage(self.content, self)
        self.pages["prikhod"]   = PrikhodPage(self.content, self)
        self.pages["raskhod"]   = RaskhodPage(self.content, self)
        self.pages["otchet"]    = OtchetPage(self.content, self)

        self.current_page = None
        self.show_page("dashboard")

    def _confirm_delete_all(self):
        if messagebox.askyesno("Удалить всё", "Удалить ВСЕ данные (приход и расход)?\nЭто действие нельзя отменить!"):
            self.delete_all()
            for page in self.pages.values():
                page.refresh()

    def show_page(self, key):
        if self.current_page:
            self.pages[self.current_page].pack_forget()
            self.nav_buttons[self.current_page].set_active(False)
        self.pages[key].pack(fill="both", expand=True)
        self.pages[key].refresh()
        self.nav_buttons[key].set_active(True)
        self.current_page = key

    def add_prikhod(self, row):
        self.data["prikhod"].append(row)
        save_data(self.data)
        self._refresh_all()

    def add_raskhod(self, row):
        self.data["raskhod"].append(row)
        save_data(self.data)
        self._refresh_all()

    def _refresh_all(self):
        for page in self.pages.values():
            page.refresh()
        self.update_idletasks()

    def delete_prikhod(self, idx):
        del self.data["prikhod"][idx]
        save_data(self.data)
        self._refresh_all()

    def delete_raskhod(self, idx):
        del self.data["raskhod"][idx]
        save_data(self.data)
        self._refresh_all()

    def delete_all(self):
        self.data = {"prikhod": [], "raskhod": []}
        save_data(self.data)
        self._refresh_all()

    def get_totals(self):
        total_in  = sum(r.get("summa", 0) for r in self.data["prikhod"])
        total_out = sum(r.get("summa", 0) for r in self.data["raskhod"])
        profit = total_out - total_in
        return total_in, total_out, profit

    def get_profit_display(self):
        ti, to, pr = self.get_totals()
        return ti, to, abs(pr), pr >= 0, "ИТОГ"

    def export_excel(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Ошибка", "Установите openpyxl:\npip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            initialfile=f"Отчёт_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            _write_sheet_prikhod(wb, self.data["prikhod"])
            _write_sheet_raskhod(wb, self.data["raskhod"])
            _write_sheet_otchet(wb, self.data)
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            wb.save(path)
            messagebox.showinfo("✅ Готово", f"Файл сохранён:\n{path}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))


# ─── КНОПКА НАВИГАЦИИ ────────────────────────────────────────────────────────
class NavButton(tk.Frame):
    def __init__(self, parent, icon, label, command):
        super().__init__(parent, bg=BG_CARD, cursor="hand2")
        self.command = command
        self.active = False
        self.icon_lbl  = tk.Label(self, text=icon,  font=("Segoe UI", 14), bg=BG_CARD, fg=TEXT_DIM, width=3)
        self.text_lbl  = tk.Label(self, text=label, font=FONT_HEAD, bg=BG_CARD, fg=TEXT_DIM, anchor="w")
        self.icon_lbl.pack(side="left", padx=(10,0), pady=10)
        self.text_lbl.pack(side="left", padx=6, pady=10, fill="x", expand=True)
        for w in (self, self.icon_lbl, self.text_lbl):
            w.bind("<Button-1>", lambda e: self.command())
            w.bind("<Enter>", self._hover_on)
            w.bind("<Leave>", self._hover_off)

    def set_active(self, val):
        self.active = val
        c = ACCENT if val else TEXT_DIM
        bg = "#1A2F3E" if val else BG_CARD
        for w in (self, self.icon_lbl, self.text_lbl):
            w.configure(bg=bg)
        self.icon_lbl.configure(fg=c)
        self.text_lbl.configure(fg=TEXT_HEADER if val else TEXT_DIM)

    def _hover_on(self, e):
        if not self.active:
            for w in (self, self.icon_lbl, self.text_lbl):
                w.configure(bg="#182A38")

    def _hover_off(self, e):
        if not self.active:
            for w in (self, self.icon_lbl, self.text_lbl):
                w.configure(bg=BG_CARD)


# ─── ОБЩИЙ ЗАГОЛОВОК СТРАНИЦЫ ────────────────────────────────────────────────
def page_header(parent, title, subtitle=""):
    fr = tk.Frame(parent, bg=BG_DARK)
    fr.pack(fill="x", padx=30, pady=(28, 0))
    tk.Label(fr, text=title, font=FONT_TITLE, bg=BG_DARK, fg=TEXT_HEADER).pack(anchor="w")
    if subtitle:
        tk.Label(fr, text=subtitle, font=FONT_SMALL, bg=BG_DARK, fg=TEXT_DIM).pack(anchor="w", pady=(2,0))
    tk.Frame(parent, bg=BORDER_CLR, height=1).pack(fill="x", padx=30, pady=14)


# ─── КАРТОЧКА СТАТИСТИКИ ─────────────────────────────────────────────────────
class StatCard(tk.Frame):
    def __init__(self, parent, icon, title, var, color=ACCENT):
        super().__init__(parent, bg=BG_CARD, padx=20, pady=16)
        tk.Label(self, text=icon, font=("Segoe UI", 22), bg=BG_CARD, fg=color).pack(anchor="w")
        self.title_lbl = tk.Label(self, text=title, font=FONT_SMALL, bg=BG_CARD, fg=TEXT_DIM)
        self.title_lbl.pack(anchor="w", pady=(4,2))
        self.val_lbl = tk.Label(self, text="0 ₸", font=FONT_BIG, bg=BG_CARD, fg=color)
        self.val_lbl.pack(anchor="w")
        self.var = var

    def update(self, val, title=None, color=None):
        self.val_lbl.configure(text=f"{val:,.0f} ₸".replace(",", " "))
        if title:
            self.title_lbl.configure(text=title)
        if color:
            self.val_lbl.configure(fg=color)


# ─── СТРАНИЦА ДАШБОРД ────────────────────────────────────────────────────────
class DashboardPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._built = False

    def _build(self):
        page_header(self, "📊 Дашборд", "Общая сводка по складу")
        cards_row = tk.Frame(self, bg=BG_DARK)
        cards_row.pack(fill="x", padx=30, pady=0)

        self.card_in     = StatCard(cards_row, "📥", "ИТОГО ПРИХОД",  "in",  ACCENT)
        self.card_out    = StatCard(cards_row, "📤", "ИТОГО РАСХОД",  "out", ACCENT2)
        self.card_profit = StatCard(cards_row, "💰", "ПРИБЫЛЬ",       "pr",  ACCENT)

        for c in (self.card_in, self.card_out, self.card_profit):
            c.pack(side="left", fill="both", expand=True, padx=6, pady=4)

        cnt_row = tk.Frame(self, bg=BG_DARK)
        cnt_row.pack(fill="x", padx=30, pady=(10, 0))
        self.cnt_in  = tk.Label(cnt_row, text="", font=FONT_SMALL, bg=BG_DARK, fg=TEXT_DIM)
        self.cnt_out = tk.Label(cnt_row, text="", font=FONT_SMALL, bg=BG_DARK, fg=TEXT_DIM)
        self.cnt_in.pack(side="left", expand=True)
        self.cnt_out.pack(side="left", expand=True)
        tk.Label(cnt_row, text="", bg=BG_DARK).pack(side="left", expand=True)

        tk.Frame(self, bg=BORDER_CLR, height=1).pack(fill="x", padx=30, pady=18)
        tk.Label(self, text="Последние операции", font=FONT_HEAD,
                 bg=BG_DARK, fg=TEXT_HEADER).pack(anchor="w", padx=30)
        self.recent_frame = tk.Frame(self, bg=BG_DARK)
        self.recent_frame.pack(fill="both", expand=True, padx=30, pady=10)

    def refresh(self):
        if not self._built:
            self._build()
            self._built = True
        ti, to, pr_raw, is_profit, pr_label = self.app.get_profit_display()
        self.card_in.update(ti)
        self.card_out.update(to)
        color = ACCENT if is_profit else RED
        self.card_profit.update(pr_raw, title=f"💰 {pr_label}", color=color)
        self.cnt_in.configure(text=f"Записей прихода: {len(self.app.data['prikhod'])}")
        self.cnt_out.configure(text=f"Записей расхода: {len(self.app.data['raskhod'])}")

        for w in self.recent_frame.winfo_children():
            w.destroy()

        all_ops = []
        for r in self.app.data["prikhod"]:
            all_ops.append(("📥 Приход", r.get("naim",""), r.get("kol",0), r.get("summa",0), r.get("data",""), ACCENT))
        for r in self.app.data["raskhod"]:
            all_ops.append(("📤 Расход", r.get("naim",""), r.get("kol",0), r.get("summa",0), r.get("data",""), ACCENT2))

        all_ops = sorted(all_ops, key=lambda x: x[4], reverse=True)[:8]
        if not all_ops:
            tk.Label(self.recent_frame, text="Пока нет данных — добавьте приход или расход",
                     font=FONT_NORM, bg=BG_DARK, fg=TEXT_DIM).pack(pady=30)
            return

        hdr = tk.Frame(self.recent_frame, bg=BG_CARD)
        hdr.pack(fill="x", pady=(0,2))
        for txt, w in [("Тип",100),("Товар",280),("Кол-во",80),("Сумма",120),("Дата",110)]:
            tk.Label(hdr, text=txt, font=FONT_SMALL, bg=BG_CARD, fg=TEXT_DIM,
                     width=w//7, anchor="w").pack(side="left", padx=8, pady=6)

        for i, (typ, naim, kol, summa, data, color) in enumerate(all_ops):
            bg = BG_ROW_ODD if i%2==0 else BG_ROW_EVEN
            row = tk.Frame(self.recent_frame, bg=bg)
            row.pack(fill="x", pady=1)
            vals = [(typ,100,color),(naim,280,TEXT_MAIN),(f"{kol} шт",80,TEXT_DIM),
                    (f"{summa:,.0f} ₸".replace(",", " "),120,TEXT_MAIN),(data,110,TEXT_DIM)]
            for txt,w,fg in vals:
                tk.Label(row, text=txt, font=FONT_SMALL, bg=bg, fg=fg,
                         width=w//7, anchor="w").pack(side="left", padx=8, pady=7)


# ─── ФОРМА ДОБАВЛЕНИЯ ────────────────────────────────────────────────────────
class AddForm(tk.Frame):
    def __init__(self, parent, fields, on_add, color=ACCENT):
        super().__init__(parent, bg=BG_CARD, padx=20, pady=16)
        self.entries = {}
        self.on_add = on_add
        self.color = color
        self.fields = fields  # сохраняем для очистки

        row = tk.Frame(self, bg=BG_CARD)
        row.pack(fill="x")

        for label, key, width in fields:
            col = tk.Frame(row, bg=BG_CARD)
            col.pack(side="left", padx=6, fill="x", expand=True)
            tk.Label(col, text=label, font=FONT_SMALL, bg=BG_CARD, fg=TEXT_DIM).pack(anchor="w")
            if key == "data":
                e = tk.Entry(col, font=FONT_NORM, bg=BG_INPUT, fg=TEXT_MAIN,
                             insertbackground=TEXT_MAIN, relief="flat",
                             width=width, bd=0, highlightthickness=1,
                             highlightbackground=BORDER_CLR, highlightcolor=color)
                e.insert(0, datetime.now().strftime("%d.%m.%Y"))
            else:
                e = tk.Entry(col, font=FONT_NORM, bg=BG_INPUT, fg=TEXT_MAIN,
                             insertbackground=TEXT_MAIN, relief="flat",
                             width=width, bd=0, highlightthickness=1,
                             highlightbackground=BORDER_CLR, highlightcolor=color)
            e.pack(fill="x", ipady=6, pady=(4,0))
            self.entries[key] = e

        # Кнопка Добавить
        btn = tk.Button(row, text="＋  Добавить", font=FONT_BTN,
                        bg=color, fg="#000",
                        relief="flat", cursor="hand2", padx=14, pady=8,
                        activebackground=ACCENT_DARK, command=self._submit)
        btn.pack(side="left", padx=(10,0), pady=(18,0))

        # ── НОВАЯ КНОПКА: Очистить поля ──────────────────────────────────────
        clear_btn = tk.Button(row, text="✕  Очистить", font=FONT_BTN,
                              bg="#2A1A1A", fg=RED,
                              relief="flat", cursor="hand2", padx=14, pady=8,
                              activebackground="#4A2020", activeforeground=RED,
                              command=self._clear)
        clear_btn.pack(side="left", padx=(6,0), pady=(18,0))

    def _clear(self):
        """Очищает все поля формы и восстанавливает дату."""
        for key, e in self.entries.items():
            e.delete(0, "end")
            if key == "data":
                e.insert(0, datetime.now().strftime("%d.%m.%Y"))

    def _submit(self):
        vals = {k: e.get().strip() for k, e in self.entries.items()}
        self.on_add(vals)


# ─── ТАБЛИЦА ─────────────────────────────────────────────────────────────────
class Table(tk.Frame):
    def __init__(self, parent, headers, color=ACCENT):
        super().__init__(parent, bg=BG_DARK)
        self.headers = headers
        self.color = color
        self._build_header()
        self.body = tk.Frame(self, bg=BG_DARK)
        self.body.pack(fill="both", expand=True)

    def _build_header(self):
        hdr = tk.Frame(self, bg=BG_CARD)
        hdr.pack(fill="x")
        for txt, w in self.headers:
            tk.Label(hdr, text=txt, font=FONT_SMALL, bg=BG_CARD,
                     fg=self.color, width=w, anchor="w").pack(side="left", padx=10, pady=8)
        tk.Label(hdr, text="", bg=BG_CARD, width=4).pack(side="left")

    def set_rows(self, rows, on_delete):
        for w in self.body.winfo_children():
            w.destroy()
        if not rows:
            tk.Label(self.body, text="Нет данных", font=FONT_NORM,
                     bg=BG_DARK, fg=TEXT_DIM).pack(pady=24)
            return
        for i, row in enumerate(rows):
            bg = BG_ROW_ODD if i%2==0 else BG_ROW_EVEN
            fr = tk.Frame(self.body, bg=bg)
            fr.pack(fill="x", pady=1)
            for (_, w), val in zip(self.headers, row):
                tk.Label(fr, text=str(val), font=FONT_SMALL, bg=bg,
                         fg=TEXT_MAIN, width=w, anchor="w").pack(side="left", padx=10, pady=7)
            idx = i
            del_btn = tk.Button(fr, text="✕", font=FONT_SMALL, bg=bg, fg=RED,
                                relief="flat", cursor="hand2", padx=4,
                                activebackground=bg,
                                command=lambda i=idx: on_delete(i))
            del_btn.pack(side="left", padx=4)


# ─── СТРАНИЦА ПРИХОД ─────────────────────────────────────────────────────────
class PrikhodPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        page_header(self, "📥 Приход товара", "Запись поступившего товара на склад")
        fields = [
            ("Наименование товара", "naim", 20),
            ("Кол-во (шт)",         "kol",  8),
            ("Цена за ед. (₸)",     "cena", 10),
            ("Поставщик",           "post", 14),
            ("Дата",                "data", 12),
        ]
        self.form = AddForm(self, fields, self._add, color=ACCENT)
        self.form.pack(fill="x", padx=30, pady=(0, 14))

        self.form.entries["kol"].bind("<KeyRelease>",  self._calc)
        self.form.entries["cena"].bind("<KeyRelease>", self._calc)

        self.sum_lbl = tk.Label(self, text="Сумма: 0 ₸", font=FONT_HEAD,
                                bg=BG_DARK, fg=ACCENT)
        self.sum_lbl.pack(anchor="w", padx=30, pady=(0,10))

        hdrs = [("№",3),("Наименование",18),("Кол-во",7),("Цена",9),("Сумма",10),("Поставщик",13),("Дата",10)]
        self.table = Table(self, hdrs, color=ACCENT)
        self.table.pack(fill="both", expand=True, padx=30, pady=4)

    def _calc(self, e=None):
        try:
            k = float(self.form.entries["kol"].get() or 0)
            c = float(self.form.entries["cena"].get() or 0)
            s = k * c
            self.sum_lbl.configure(text=f"Сумма: {s:,.0f} ₸".replace(",", " "))
        except:
            self.sum_lbl.configure(text="Сумма: —")

    def _add(self, vals):
        if not vals["naim"]:
            messagebox.showwarning("Внимание", "Введите наименование товара")
            return
        try:
            kol  = float(vals["kol"] or 0)
            cena = float(vals["cena"] or 0)
        except:
            messagebox.showerror("Ошибка", "Кол-во и цена должны быть числами")
            return
        summa = kol * cena
        rec = {"naim": vals["naim"], "kol": kol, "cena": cena,
               "summa": summa, "post": vals["post"], "data": vals["data"]}
        self.app.add_prikhod(rec)
        for key, e in self.form.entries.items():
            e.delete(0, "end")
            if key == "data":
                e.insert(0, datetime.now().strftime("%d.%m.%Y"))
        self.sum_lbl.configure(text="Сумма: 0 ₸")
        self.refresh()

    def _delete(self, idx):
        if messagebox.askyesno("Удалить", "Удалить эту запись?"):
            self.app.delete_prikhod(idx)
            self.refresh()

    def refresh(self):
        if not self.app.data["prikhod"]:
            self.sum_lbl.configure(text="Сумма: 0 ₸")
        rows = []
        for i, r in enumerate(self.app.data["prikhod"], 1):
            rows.append([i, r.get("naim",""), int(r.get("kol",0)),
                         f"{r.get('cena',0):,.0f}".replace(",", " "),
                         f"{r.get('summa',0):,.0f}".replace(",", " "),
                         r.get("post",""), r.get("data","")])
        self.table.set_rows(rows, self._delete)


# ─── СТРАНИЦА РАСХОД ─────────────────────────────────────────────────────────
class RaskhodPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        page_header(self, "📤 Расход товара", "Запись проданного товара")
        fields = [
            ("Наименование товара", "naim", 20),
            ("Кол-во (шт)",         "kol",  8),
            ("Цена продажи (₸)",    "cena", 10),
            ("Покупатель",          "buyer",14),
            ("Дата",                "data", 12),
        ]
        self.form = AddForm(self, fields, self._add, color=ACCENT2)
        self.form.pack(fill="x", padx=30, pady=(0,14))

        self.form.entries["kol"].bind("<KeyRelease>",  self._calc)
        self.form.entries["cena"].bind("<KeyRelease>", self._calc)

        self.sum_lbl = tk.Label(self, text="Сумма: 0 ₸", font=FONT_HEAD,
                                bg=BG_DARK, fg=ACCENT2)
        self.sum_lbl.pack(anchor="w", padx=30, pady=(0,10))

        hdrs = [("№",3),("Наименование",18),("Кол-во",7),("Цена",9),("Сумма",10),("Покупатель",13),("Дата",10)]
        self.table = Table(self, hdrs, color=ACCENT2)
        self.table.pack(fill="both", expand=True, padx=30, pady=4)

    def _calc(self, e=None):
        try:
            k = float(self.form.entries["kol"].get() or 0)
            c = float(self.form.entries["cena"].get() or 0)
            s = k * c
            self.sum_lbl.configure(text=f"Сумма: {s:,.0f} ₸".replace(",", " "))
        except:
            self.sum_lbl.configure(text="Сумма: —")

    def _add(self, vals):
        if not vals["naim"]:
            messagebox.showwarning("Внимание", "Введите наименование товара")
            return
        try:
            kol  = float(vals["kol"] or 0)
            cena = float(vals["cena"] or 0)
        except:
            messagebox.showerror("Ошибка", "Кол-во и цена должны быть числами")
            return
        summa = kol * cena
        rec = {"naim": vals["naim"], "kol": kol, "cena": cena,
               "summa": summa, "buyer": vals["buyer"], "data": vals["data"]}
        self.app.add_raskhod(rec)
        for key, e in self.form.entries.items():
            e.delete(0, "end")
            if key == "data":
                e.insert(0, datetime.now().strftime("%d.%m.%Y"))
        self.sum_lbl.configure(text="Сумма: 0 ₸")
        self.refresh()

    def _delete(self, idx):
        if messagebox.askyesno("Удалить", "Удалить эту запись?"):
            self.app.delete_raskhod(idx)
            self.refresh()

    def refresh(self):
        if not self.app.data["raskhod"]:
            self.sum_lbl.configure(text="Сумма: 0 ₸")
        rows = []
        for i, r in enumerate(self.app.data["raskhod"], 1):
            rows.append([i, r.get("naim",""), int(r.get("kol",0)),
                         f"{r.get('cena',0):,.0f}".replace(",", " "),
                         f"{r.get('summa',0):,.0f}".replace(",", " "),
                         r.get("buyer",""), r.get("data","")])
        self.table.set_rows(rows, self._delete)


# ─── СТРАНИЦА ОТЧЁТ ──────────────────────────────────────────────────────────
class OtchetPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG_DARK)
        self.app = app
        self._build()

    def _build(self):
        page_header(self, "📋 Итоговый отчёт", "Сводная таблица по всем операциям")

        self.summary = tk.Frame(self, bg=BG_DARK)
        self.summary.pack(fill="x", padx=30, pady=(0,18))
        self.lbl_in     = self._sum_box(self.summary, "📥 Общий приход", ACCENT)
        self.lbl_out    = self._sum_box(self.summary, "📤 Общий расход", ACCENT2)
        self.lbl_profit = self._sum_box(self.summary, "💰 Прибыль",      ACCENT)

        tk.Label(self, text="Сводка по наименованиям", font=FONT_HEAD,
                 bg=BG_DARK, fg=TEXT_HEADER).pack(anchor="w", padx=30, pady=(0,6))
        hdrs = [("Наименование",20),("Приход шт",8),("Приход ₸",10),
                ("Расход шт",8),("Расход ₸",10),("Остаток",8),("Разница ₸",11)]
        self.table = Table(self, hdrs, color=ACCENT)
        self.table.pack(fill="both", expand=True, padx=30, pady=4)

    def _sum_box(self, parent, title, color):
        fr = tk.Frame(parent, bg=BG_CARD, padx=16, pady=12)
        fr.pack(side="left", fill="both", expand=True, padx=6, pady=4)
        tk.Label(fr, text=title, font=FONT_SMALL, bg=BG_CARD, fg=TEXT_DIM).pack(anchor="w")
        lbl = tk.Label(fr, text="0 ₸", font=("Segoe UI", 18, "bold"), bg=BG_CARD, fg=color)
        lbl.pack(anchor="w", pady=(4,0))
        return lbl

    def refresh(self):
        ti, to, pr_abs, is_profit, pr_label = self.app.get_profit_display()
        self.lbl_in.configure(text=f"{ti:,.0f} ₸".replace(",", " "))
        self.lbl_out.configure(text=f"{to:,.0f} ₸".replace(",", " "))
        color = ACCENT if is_profit else RED
        self.lbl_profit.configure(fg=color, text=f"{pr_abs:,.0f} ₸".replace(",", " "))

        merged = {}
        for r in self.app.data["prikhod"]:
            n = r.get("naim","")
            if n not in merged:
                merged[n] = {"pi":0,"ps":0,"ri":0,"rs":0}
            merged[n]["pi"] += r.get("kol",0)
            merged[n]["ps"] += r.get("summa",0)
        for r in self.app.data["raskhod"]:
            n = r.get("naim","")
            if n not in merged:
                merged[n] = {"pi":0,"ps":0,"ri":0,"rs":0}
            merged[n]["ri"] += r.get("kol",0)
            merged[n]["rs"] += r.get("summa",0)

        rows = []
        for name, v in merged.items():
            ostatok = v["pi"] - v["ri"]
            raznica = abs(v["rs"] - v["ps"])
            rows.append([name, int(v["pi"]),
                         f"{v['ps']:,.0f}".replace(",", " "),
                         int(v["ri"]),
                         f"{v['rs']:,.0f}".replace(",", " "),
                         int(ostatok),
                         f"{raznica:,.0f}".replace(",", " ")])
        self.table.set_rows(rows, lambda i: None)


# ─── ЭКСПОРТ В EXCEL ─────────────────────────────────────────────────────────
def _xl_style(ws, row, col, value, bold=False, bg=None, fg=None, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg or "000000", name="Segoe UI", size=10)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    return cell

def _write_sheet_prikhod(wb, data):
    ws = wb.create_sheet("ПРИХОД")
    ws.row_dimensions[1].height = 30
    headers = ["№","Наименование","Кол-во","Цена (₸)","Сумма (₸)","Поставщик","Дата"]
    for c, h in enumerate(headers, 1):
        _xl_style(ws, 1, c, h, bold=True, bg="1B3A4B", fg="00C896", align="center")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 14
    for i, r in enumerate(data, 2):
        ws.row_dimensions[i].height = 20
        bg = "1C2E3D" if i%2==0 else "16232E"
        _xl_style(ws, i, 1, i-1,              bg=bg, fg="FFFFFF", align="center")
        _xl_style(ws, i, 2, r.get("naim",""), bg=bg, fg="FFFFFF")
        _xl_style(ws, i, 3, r.get("kol",0),   bg=bg, fg="E8F4F0", align="center")
        _xl_style(ws, i, 4, r.get("cena",0),  bg=bg, fg="E8F4F0", align="right")
        _xl_style(ws, i, 5, r.get("summa",0), bg=bg, fg="00C896", align="right")
        _xl_style(ws, i, 6, r.get("post",""), bg=bg, fg="E8F4F0")
        _xl_style(ws, i, 7, r.get("data",""), bg=bg, fg="7A9FAF", align="center")
    last = len(data) + 2
    total = sum(r.get("summa",0) for r in data)
    _xl_style(ws, last, 1, "ИТОГО:", bold=True, bg="0A1F2E", fg="FFFFFF")
    _xl_style(ws, last, 5, total,   bold=True, bg="0A1F2E", fg="00C896", align="right")

def _write_sheet_raskhod(wb, data):
    ws = wb.create_sheet("РАСХОД")
    ws.row_dimensions[1].height = 30
    headers = ["№","Наименование","Кол-во","Цена (₸)","Сумма (₸)","Покупатель","Дата"]
    for c, h in enumerate(headers, 1):
        _xl_style(ws, 1, c, h, bold=True, bg="2E1E00", fg="F0A500", align="center")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 14
    for i, r in enumerate(data, 2):
        ws.row_dimensions[i].height = 20
        bg = "221A00" if i%2==0 else "1A1400"
        _xl_style(ws, i, 1, i-1,               bg=bg, fg="FFFFFF", align="center")
        _xl_style(ws, i, 2, r.get("naim",""),  bg=bg, fg="FFFFFF")
        _xl_style(ws, i, 3, r.get("kol",0),    bg=bg, fg="E8F4F0", align="center")
        _xl_style(ws, i, 4, r.get("cena",0),   bg=bg, fg="E8F4F0", align="right")
        _xl_style(ws, i, 5, r.get("summa",0),  bg=bg, fg="F0A500", align="right")
        _xl_style(ws, i, 6, r.get("buyer",""), bg=bg, fg="E8F4F0")
        _xl_style(ws, i, 7, r.get("data",""),  bg=bg, fg="7A9FAF", align="center")
    last = len(data) + 2
    total = sum(r.get("summa",0) for r in data)
    _xl_style(ws, last, 1, "ИТОГО:", bold=True, bg="0A1F2E", fg="FFFFFF")
    _xl_style(ws, last, 5, total,   bold=True, bg="0A1F2E", fg="F0A500", align="right")

def _write_sheet_otchet(wb, data):
    ws = wb.create_sheet("ИТОГОВЫЙ ОТЧЁТ")
    ws.row_dimensions[1].height = 30
    headers = ["Наименование","Приход шт","Приход ₸","Расход шт","Расход ₸","Остаток шт","Разница ₸"]
    for c, h in enumerate(headers, 1):
        _xl_style(ws, 1, c, h, bold=True, bg="0A1F2E", fg="00C896", align="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["G"].width = 16

    merged = {}
    for r in data["prikhod"]:
        n = r.get("naim","")
        if n not in merged: merged[n] = {"pi":0,"ps":0,"ri":0,"rs":0}
        merged[n]["pi"] += r.get("kol",0)
        merged[n]["ps"] += r.get("summa",0)
    for r in data["raskhod"]:
        n = r.get("naim","")
        if n not in merged: merged[n] = {"pi":0,"ps":0,"ri":0,"rs":0}
        merged[n]["ri"] += r.get("kol",0)
        merged[n]["rs"] += r.get("summa",0)

    for i, (name, v) in enumerate(merged.items(), 2):
        ws.row_dimensions[i].height = 20
        bg = "1C2E3D" if i%2==0 else "16232E"
        ostatok = v["pi"] - v["ri"]
        raznica = v["rs"] - v["ps"]
        raznica_abs = abs(raznica)
        raznica_color = "00C896" if raznica >= 0 else "FF5A5A"
        _xl_style(ws, i, 1, name,        bg=bg, fg="FFFFFF")
        _xl_style(ws, i, 2, int(v["pi"]),bg=bg, fg="E8F4F0", align="center")
        _xl_style(ws, i, 3, v["ps"],     bg=bg, fg="00C896", align="right")
        _xl_style(ws, i, 4, int(v["ri"]),bg=bg, fg="E8F4F0", align="center")
        _xl_style(ws, i, 5, v["rs"],     bg=bg, fg="F0A500", align="right")
        _xl_style(ws, i, 6, int(ostatok),bg=bg, fg="E8F4F0", align="center")
        _xl_style(ws, i, 7, raznica_abs, bg=bg, fg=raznica_color, align="right")

    last = len(merged) + 2
    ws.row_dimensions[last].height = 24
    ti = sum(v["ps"] for v in merged.values())
    to = sum(v["rs"] for v in merged.values())
    pr = to - ti
    pr_abs = abs(pr)
    _xl_style(ws, last, 1, "ИТОГО:",  bold=True, bg="0A1F2E", fg="FFFFFF")
    _xl_style(ws, last, 3, ti,        bold=True, bg="0A1F2E", fg="00C896", align="right")
    _xl_style(ws, last, 5, to,        bold=True, bg="0A1F2E", fg="F0A500", align="right")
    _xl_style(ws, last, 7, pr_abs,    bold=True, bg="0A1F2E", fg="00C896" if pr>=0 else "FF5A5A", align="right")


# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
